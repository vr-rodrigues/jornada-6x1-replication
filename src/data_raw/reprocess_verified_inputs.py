"""Rebuild PNAD 2024Q4 / RAIS 2022 aggregates from Base dos Dados BigQuery.

Only aggregate cells leave BigQuery. Originals are never read as fallback or
overwritten. Every query gets a dry run, a byte cap, saved SQL, and job metadata.
Run: python src/data_raw/reprocess_verified_inputs.py --project upa-research
Use --from-cache to recompute from this collector's archived aggregate cells.
"""
from __future__ import annotations
import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data_intermediate/reprocessed"
FINAL = ROOT / "data_final/reprocessed"
PROV = OUT / "provenance"

PNAD_SQL = """
WITH source AS (
 SELECT V1028 w, V2009 age, SAFE_CAST(VD4002 AS INT64) occupied,
 SAFE_CAST(VD4009 AS INT64) position, SAFE_CAST(V4019 AS INT64) cnpj,
 SAFE_CAST(V4017 AS INT64) partner, SAFE_CAST(V4018 AS INT64) size_category,
 V40183 size_exact_11_50, V4039 habitual, V4039C actual,
 VD4031 habitual_all_jobs, VD4035 actual_all_jobs, VD4016 income_monthly,
 SAFE_CAST(SUBSTR(LPAD(V4013,5,'0'),1,2) AS INT64) cnae2
 FROM `basedosdados.br_ibge_pnadc.microdados`
 WHERE ano=2024 AND trimestre=4
), classified AS (
 SELECT *, CASE WHEN cnae2 BETWEEN 1 AND 3 THEN 'agriculture'
 WHEN cnae2 BETWEEN 5 AND 43 THEN 'industry'
 WHEN cnae2 BETWEEN 45 AND 99 THEN 'services' ELSE 'unclassified' END sector,
 CASE WHEN position IN (2,4,10) THEN 1
 WHEN position IN (1,3,5,6,7) THEN 0
 WHEN position IN (8,9) AND cnpj=1 THEN 0
 WHEN position IN (8,9) AND cnpj=2 THEN 1 ELSE NULL END informal,
 CASE WHEN size_category IN (1,2) OR (size_category=3 AND size_exact_11_50<50) THEN 'le49'
 WHEN size_category=3 AND size_exact_11_50=50 THEN 'eq50'
 WHEN size_category=4 THEN 'ge51' ELSE 'unknown' END size_group
 FROM source WHERE age>=14 AND occupied=1
)
SELECT sector, position, cnpj, partner, informal, size_group, habitual,
 COUNT(*) sample_n, COUNTIF(w IS NULL OR w<=0) invalid_weight_n,
 SUM(IF(w>0,w,0)) occupied_weighted,
 COUNTIF(actual IS NULL) missing_actual_n,
 COUNTIF(habitual IS NULL) missing_habitual_n,
 COUNTIF(income_monthly IS NULL) missing_income_n,
 SUM(IF(w>0 AND actual BETWEEN 0 AND 120,w,0)) actual_weight,
 SUM(IF(w>0 AND actual BETWEEN 0 AND 120,w*actual,0)) actual_hours_sum,
 SUM(IF(w>0 AND habitual_all_jobs BETWEEN 1 AND 120,w,0)) habitual_all_weight,
 SUM(IF(w>0 AND habitual_all_jobs BETWEEN 1 AND 120,w*habitual_all_jobs,0)) habitual_all_sum,
 SUM(IF(w>0 AND actual_all_jobs BETWEEN 0 AND 120,w,0)) actual_all_weight,
 SUM(IF(w>0 AND actual_all_jobs BETWEEN 0 AND 120,w*actual_all_jobs,0)) actual_all_sum,
 SUM(IF(w>0 AND habitual BETWEEN 1 AND 120 AND income_monthly>0,w,0)) paid_weight,
 SUM(IF(w>0 AND habitual BETWEEN 1 AND 120 AND income_monthly>0,w*income_monthly,0)) paid_income_monthly,
 SUM(IF(w>0 AND habitual BETWEEN 1 AND 120 AND income_monthly>0,w*habitual,0)) paid_hours,
 SUM(IF(w>0 AND habitual BETWEEN 1 AND 120 AND income_monthly>0,w*income_monthly/(habitual*52.0/12.0),0)) paid_individual_hourly_sum
FROM classified GROUP BY sector,position,cnpj,partner,informal,size_group,habitual
ORDER BY sector,position,cnpj,partner,informal,size_group,habitual
"""
PNAD_POP_SQL = """
SELECT ano,trimestre,COUNT(*) sample_n,
COUNTIF(V2009>=14) working_age_sample_n,
SUM(IF(V1028>0,V1028,0)) population_weighted,
SUM(IF(V2009>=14 AND V1028>0,V1028,0)) working_age_weighted,
COUNTIF(V1028 IS NULL OR V1028<=0) invalid_weight_n
FROM `basedosdados.br_ibge_pnadc.microdados`
WHERE ano=2024 AND trimestre=4 GROUP BY ano,trimestre
"""
RAIS_SQL = """
SELECT tamanho_estabelecimento, vinculo_ativo_3112,
quantidade_horas_contratadas, COUNT(*) n_vinculos
FROM `basedosdados.br_me_rais.microdados_vinculos`
WHERE ano=2022
GROUP BY tamanho_estabelecimento,vinculo_ativo_3112,quantidade_horas_contratadas
ORDER BY tamanho_estabelecimento,vinculo_ativo_3112,quantidade_horas_contratadas
"""
DICTS = {
 "pnad_dictionary": "SELECT * FROM `basedosdados.br_ibge_pnadc.dicionario` WHERE UPPER(nome_coluna) IN ('VD4002','VD4009','V4019','V4017','V4018','V40183','V4039','V4039C','V1028') ORDER BY nome_coluna,chave",
 "rais_dictionary": "SELECT * FROM `basedosdados.br_me_rais.dicionario` WHERE nome_coluna IN ('tamanho_estabelecimento','vinculo_ativo_3112','quantidade_horas_contratadas') ORDER BY nome_coluna,chave",
}

def save(path, value):
 path.parent.mkdir(parents=True,exist_ok=True)
 path.write_text(json.dumps(value,ensure_ascii=False,indent=2,allow_nan=False),encoding="utf-8")

def sha256_file(path):
 digest=hashlib.sha256()
 with path.open("rb") as f:
  for chunk in iter(lambda:f.read(1024*1024),b""):digest.update(chunk)
 return digest.hexdigest()

def ratio(a,b):
 return a/b if a is not None and b else None

def classify_informality(position,cnpj):
 """IBGE classification; missing CNPJ is unknown, never automatically no."""
 if position in (2,4,10):return 1
 if position in (1,3,5,6,7):return 0
 if position in (8,9) and cnpj in (1,2):return int(cnpj==2)
 return None

def query(client,name,sql,cap):
 from google.cloud import bigquery
 (PROV/f"{name}.sql").write_text(sql,encoding="utf-8")
 dry=client.query(sql,job_config=bigquery.QueryJobConfig(dry_run=True,use_query_cache=False))
 estimated=int(dry.total_bytes_processed or 0)
 print(f"{name}: dry-run {estimated:,} bytes; cap {cap:,}",flush=True)
 if estimated>cap:
  raise RuntimeError(f"{name}: estimated {estimated} exceeds explicit byte cap {cap}")
 job=client.query(sql,job_config=bigquery.QueryJobConfig(maximum_bytes_billed=cap,use_query_cache=False,labels={"study":"jornada-replication"}))
 rows=[dict(r.items()) for r in job.result()]
 metadata={"job_id":job.job_id,"project":job.project,"location":job.location,
 "created":str(job.created),"ended":str(job.ended),"estimated_bytes":estimated,
 "total_bytes_processed":job.total_bytes_processed,"total_bytes_billed":job.total_bytes_billed,
 "cache_hit":job.cache_hit,"maximum_bytes_billed":cap,
 "sql_sha256":hashlib.sha256(sql.encode()).hexdigest(),"row_count":len(rows)}
 save(PROV/f"{name}_job.json",metadata)
 save(PROV/f"{name}_rows.json",rows)
 print(f"{name}: {len(rows)} aggregate rows, job={job.job_id}",flush=True)
 return rows

def summarize(rows):
 total=sum(r["occupied_weighted"] for r in rows)
 known=sum(r["occupied_weighted"] for r in rows if r["informal"] in (0,1))
 informal=sum(r["occupied_weighted"] for r in rows if r["informal"]==1)
 d={"occupied_weighted":total,"sample_n":sum(r["sample_n"] for r in rows),
 "informality_rate":ratio(informal,known),"informality_denominator_weighted":known,
 "unknown_informality_weighted":total-known,
 "informality_lower_bound_all_occupied":ratio(informal,total),
 "informality_upper_bound_all_occupied":ratio(informal+total-known,total),
 "mean_hours_habitual":{},"mean_hours_actual":{},"mean_hours_habitual_all_jobs":{},"mean_hours_actual_all_jobs":{},"wages":{},
 "contracted_hours":None,"contracted_hours_reason":"PNAD measures habitual/actual hours; no contracted-hours field used."}
 for label,status in [("total",None),("formal",0),("informal",1)]:
  group=rows if status is None else [r for r in rows if r["informal"]==status]
  valid=[r for r in group if r["habitual"] is not None and 1<=r["habitual"]<=120]
  weight=sum(r["occupied_weighted"] for r in valid)
  d["mean_hours_habitual"][label]=ratio(sum(r["occupied_weighted"]*r["habitual"] for r in valid),weight)
  d["mean_hours_actual"][label]=ratio(sum(r["actual_hours_sum"] for r in group),sum(r["actual_weight"] for r in group))
  d["mean_hours_habitual_all_jobs"][label]=ratio(sum(r["habitual_all_sum"] for r in group),sum(r["habitual_all_weight"] for r in group))
  d["mean_hours_actual_all_jobs"][label]=ratio(sum(r["actual_all_sum"] for r in group),sum(r["actual_all_weight"] for r in group))
  pw=sum(r["paid_weight"] for r in group); pay=sum(r["paid_income_monthly"] for r in group); hours=sum(r["paid_hours"] for r in group)
  hours_capped44=sum(min(r["habitual"],44)*r["paid_weight"] for r in group if r["habitual"] is not None and r["paid_weight"]>0)
  d["wages"][label]={"paid_workers_weighted":pw,"income_monthly_sum":pay,"hours_weekly_sum":hours,
   "hours_weekly_sum_capped44":hours_capped44,"aggregate_hourly_payroll_over_hours_capped44":ratio(pay*12/52,hours_capped44),
   "mean_monthly_per_worker":ratio(pay,pw),"mean_weekly_per_worker":ratio(pay*12/52,pw),
   "aggregate_hourly_payroll_over_hours":ratio(pay*12/52,hours),
   "mean_individual_hourly":ratio(sum(r["paid_individual_hourly_sum"] for r in group),pw),
   "hours_valid_weighted":weight,"hours_missing_or_invalid_weighted":sum(r["occupied_weighted"] for r in group)-weight,
   "wage_bridge_excluded_weighted":sum(r["occupied_weighted"] for r in group)-pw}
 hours=defaultdict(float)
 for r in rows:
  if r["informal"]==0 and r["habitual"] is not None and 1<=r["habitual"]<=120:
   hours[r["habitual"]]+=r["occupied_weighted"]
 denom=sum(hours.values())
 d["formal_hours_distribution"]={"hours":sorted(hours),"weights":[hours[h]/denom for h in sorted(hours)],"measure":"habitual_main_job","denominator_weighted":denom}
 d["formal_hours_proxy_cap44"]={"share_above44":ratio(sum(w for h,w in hours.items() if h>44),denom),"mean_habitual_capped44":ratio(sum(min(h,44)*w for h,w in hours.items()),denom),"assumption":"Applying a legal cap to reported habitual hours is a model proxy, not observed contracted hours."}
 d["formal_hours_bins_comparable"]={"theta_36":ratio(sum(w for h,w in hours.items() if h<=36),denom),"theta_40":ratio(sum(w for h,w in hours.items() if 36<h<=40),denom),"theta_44":ratio(sum(w for h,w in hours.items() if h>40),denom),"definition":"<=36, 37-40, >40 habitual hours; representative 36/40/44 is a separate model approximation"}
 d["wage_ratio_formal_informal"]={key:ratio(d["wages"]["formal"][key],d["wages"]["informal"][key]) for key in ["aggregate_hourly_payroll_over_hours","mean_monthly_per_worker","mean_weekly_per_worker","mean_individual_hourly"]}
 d["wage_ratio_formal_informal"]["aggregate_hourly_formal_capped44"]=ratio(d["wages"]["formal"]["aggregate_hourly_payroll_over_hours_capped44"],d["wages"]["informal"]["aggregate_hourly_payroll_over_hours"])
 return d

def build_pnad(rows,pop):
 if len(pop)!=1 or pop[0]["ano"]!=2024 or pop[0]["trimestre"]!=4 or not rows:
  raise ValueError("PNAD 2024Q4 is unavailable; refusing a quarter substitution")
 national=summarize(rows)
 sectors={s:summarize([r for r in rows if r["sector"]==s]) for s in sorted({r["sector"] for r in rows})}
 for d in sectors.values(): d["employment_share"]=d["occupied_weighted"]/national["occupied_weighted"]
 national["employment_share"]=1.0
 national["employment_to_population_14plus"]=national["occupied_weighted"]/pop[0]["working_age_weighted"]
 national["size_groups"]={s:summarize([r for r in rows if r["size_group"]==s]) for s in sorted({r["size_group"] for r in rows})}
 for label,d in [("national",national),*sectors.items()]:
  scope=rows if label=="national" else [r for r in rows if r["sector"]==label]
  employee_summary=summarize([r for r in scope if r["position"] in (1,2)])
  d["private_employee_wage_bridge"]={"universe":"Private-sector employees VD4009=1 or2, positive habitual main-job earnings and valid habitual hours; excludes employer/self-employed income","wage_ratio_formal_informal":employee_summary["wage_ratio_formal_informal"],"wages":employee_summary["wages"]}
 result={"status":"verified_reprocessed","metadata":{"source":"IBGE PNAD Continua via Base dos Dados BigQuery","table":"basedosdados.br_ibge_pnadc.microdados","year":2024,"quarter":4,"weights":"V1028","universe":"Persons aged 14+ occupied in reference week (VD4002=1); main-job characteristics; all Brazil","cnpj":"V4019: 1 registered, 2 not registered, other/missing unknown for employer/self-employed","hours":"V4039 habitual main job; V4039C actual main job; contracted hours unavailable","wage_bridge":"Positive VD4016 habitual monthly income and valid habitual main-job hours, same observations in payroll and hours denominators; monthly-to-weekly factor 12/52; formal/informal ratio","uncertainty":"Point estimates using survey weights; no design-based standard errors; no claim of identified structural parameters","retrieved_utc":dt.datetime.now(dt.timezone.utc).isoformat()},"national":national,"sectors":sectors,"quality":{"population":pop[0],"invalid_weight_n":sum(r["invalid_weight_n"] for r in rows),"unknown_informality_n":sum(r["sample_n"] for r in rows if r["informal"] is None),"missing_habitual_n":sum(r["missing_habitual_n"] for r in rows),"missing_actual_n":sum(r["missing_actual_n"] for r in rows),"missing_income_n":sum(r["missing_income_n"] for r in rows)}}
 save(OUT/"pnad_targets.json",result)
 save(FINAL/"pnad_targets.json",result)
 flat=[]
 for name,d in [("NATIONAL",national),*sectors.items()]:
  flat.append({"sector":name,"lambda_s":d["employment_share"],"inf_rate":d["informality_rate"],"occupied_weighted":d["occupied_weighted"],"sample_n":d["sample_n"],**{k:v for k,v in d["formal_hours_bins_comparable"].items() if k.startswith("theta_")},"avg_hours_formal_habitual":d["mean_hours_habitual"]["formal"],"avg_hours_formal_actual":d["mean_hours_actual"]["formal"],"avg_hours_informal_habitual":d["mean_hours_habitual"]["informal"],"wage_ratio_hourly_payroll":d["wage_ratio_formal_informal"]["aggregate_hourly_payroll_over_hours"]})
 with (FINAL/"SECTORAL_PNAD_EMPIRICAL.csv").open("w",newline="",encoding="utf-8") as f:
  w=csv.DictWriter(f,fieldnames=flat[0]);w.writeheader();w.writerows(flat)
 return result

def build_rais(rows,dictionary):
 labels={r["chave"]:r["valor"] for r in dictionary if r["nome_coluna"]=="tamanho_estabelecimento"}
 active=[r for r in rows if str(r["vinculo_ativo_3112"])=="1"]
 by_size=defaultdict(int)
 for r in active: by_size[str(r["tamanho_estabelecimento"])]+=r["n_vinculos"]
 # BD dictionary codes: 1=zero, 2=1..4, 3=5..9, 4=10..19, 5=20..49.
 for code in ["2","3","4","5"]:
  if code not in labels: raise ValueError("Missing RAIS size dictionary; refusing an assumed code mapping")
 total=sum(by_size.values()); small=sum(by_size.get(k,0) for k in ["2","3","4","5"])
 result={"status":"verified_reprocessed","source":"MTE RAIS via Base dos Dados BigQuery","table":"basedosdados.br_me_rais.microdados_vinculos","year":2022,"universe":"Employment links active 31 December 2022, all Brazil and legal employment types; administrative counts, no survey weights","unit":"employment links (not unique people)","size_unit":"establishment (not consolidated company)","n_active_links":total,"small_le49_share_formal":small/total,"large_ge50_share_formal":sum(v for k,v in by_size.items() if k in ["6","7","8","9","10"])/total,"zero_or_unknown_size_active_links":total-small-sum(v for k,v in by_size.items() if k in ["6","7","8","9","10"]),"breakdown":[{"code":k,"label":labels.get(k),"n_active_links":v,"share":v/total} for k,v in sorted(by_size.items())],"original_claims":{"small_share":0.59,"subcategories_sum":0.32,"status":"Neither original percentage is accepted merely because it appears in the legacy JSON; compare these fresh administrative counts."},"pnad_comparability":"PNAD counts persons, includes registered self-employed/employers and household workers; do not combine these RAIS shares with PNAD informal rates without a universe bridge."}
 hours=defaultdict(int)
 for r in active: hours[str(r["quantidade_horas_contratadas"])]+=r["n_vinculos"]
 result["contracted_hours_counts"]=[{"hours":k,"n_active_links":v} for k,v in sorted(hours.items())]
 save(OUT/"rais_targets.json",result);save(FINAL/"rais_targets.json",result)
 return result

IBGE_BASE="https://ftp.ibge.gov.br/Trabalho_e_Rendimento/Pesquisa_Nacional_por_Amostra_de_Domicilios_continua/Trimestral/Microdados/"
IBGE_FILE="PNADC_042024_20250815.zip"
MTE_BASE="https://www.gov.br/trabalho-e-emprego/pt-br/acesso-a-informacao/acoes-e-programas/programas-projetos-acoes-obras-e-atividades/estatisticas-trabalho/rais/rais-2022/"

def download(url,path):
 import requests
 if not path.exists():
  path.parent.mkdir(parents=True,exist_ok=True)
  r=requests.get(url,stream=True,timeout=90);r.raise_for_status()
  tmp=path.with_suffix(path.suffix+".partial")
  with tmp.open("wb") as f:
   for chunk in r.iter_content(1024*1024):
    if chunk:f.write(chunk)
  tmp.replace(path)
 digest=sha256_file(path)
 save(PROV/(path.name+".source.json"),{"url":url,"path":str(path.relative_to(ROOT)),"sha256":digest,"bytes":path.stat().st_size,"retrieved_or_verified_utc":dt.datetime.now(dt.timezone.utc).isoformat()})
 return path

def official_pnad():
 """Explicit fallback only: official 2024Q4 vintage, never another quarter."""
 micro=download(IBGE_BASE+"2024/"+IBGE_FILE,ROOT/"data_raw/reprocessed/ibge"/IBGE_FILE)
 dictionary=download(IBGE_BASE+"Documentacao/Dicionario_e_input_20221031.zip",PROV/"Dicionario_e_input_20221031.zip")
 with zipfile.ZipFile(dictionary) as z:
  sas=z.read("input_PNADC_trimestral.txt").decode("latin1")
  (PROV/"input_PNADC_trimestral.txt").write_text(sas,encoding="utf-8")
  (PROV/"dicionario_PNADC_microdados_trimestral.xls").write_bytes(z.read("dicionario_PNADC_microdados_trimestral.xls"))
 specs={name:(int(start)-1,int(width)) for start,name,width in re.findall(r"@(\d+)\s+(\w+)\s+\$?(\d+)\.",sas)}
 def value(line,name):
  start,width=specs[name];raw=line[start:start+width].strip()
  return float(raw) if raw else None
 sums=["sample_n","invalid_weight_n","occupied_weighted","missing_actual_n","missing_habitual_n","missing_income_n","actual_weight","actual_hours_sum","habitual_all_weight","habitual_all_sum","actual_all_weight","actual_all_sum","paid_weight","paid_income_monthly","paid_hours","paid_individual_hourly_sum"]
 grouped={}
 population={"ano":2024,"trimestre":4,"sample_n":0,"working_age_sample_n":0,"population_weighted":0.,"working_age_weighted":0.,"invalid_weight_n":0}
 with zipfile.ZipFile(micro) as z:
  names=z.namelist()
  if names!=["PNADC_042024.txt"]:raise ValueError("Unexpected microdata archive contents")
  with z.open(names[0]) as f:
   for line in f:
    if line[:5]!=b"20244":raise ValueError("Unexpected quarter in official archive")
    w=value(line,"V1028");age=value(line,"V2009")
    population["sample_n"]+=1
    if w is None or w<=0:population["invalid_weight_n"]+=1
    w=max(w or 0,0)
    population["population_weighted"]+=w
    if age is None or age<14:continue
    population["working_age_sample_n"]+=1;population["working_age_weighted"]+=w
    if value(line,"VD4002")!=1:continue
    pos=value(line,"VD4009");cnpj=value(line,"V4019");partner=value(line,"V4017")
    inf=classify_informality(pos,cnpj)
    cnae=value(line,"V4013")
    # CNAE Domiciliar 2.0 has five digits; the first two identify its division.
    cnae2=int(str(int(cnae)).zfill(5)[:2]) if cnae is not None else None
    sector="agriculture" if cnae2 is not None and 1<=cnae2<=3 else "industry" if cnae2 is not None and 5<=cnae2<=43 else "services" if cnae2 is not None and 45<=cnae2<=99 else "unclassified"
    sz=value(line,"V4018");exact=value(line,"V40183")
    size="le49" if sz in (1,2) or (sz==3 and exact is not None and exact<50) else "eq50" if sz==3 and exact==50 else "ge51" if sz==4 else "unknown"
    h=value(line,"V4039");a=value(line,"V4039C");ha=value(line,"VD4031");aa=value(line,"VD4035");inc=value(line,"VD4016")
    key=(sector,pos,cnpj,partner,inf,size,h)
    if key not in grouped:grouped[key]={**dict(zip(["sector","position","cnpj","partner","informal","size_group","habitual"],key)),**dict.fromkeys(sums,0)}
    g=grouped[key];g["sample_n"]+=1;g["occupied_weighted"]+=w;g["invalid_weight_n"]+=int(w<=0)
    g["missing_actual_n"]+=int(a is None);g["missing_habitual_n"]+=int(h is None);g["missing_income_n"]+=int(inc is None)
    for measure,label,minimum in [(a,"actual",0),(ha,"habitual_all",1),(aa,"actual_all",0)]:
     if measure is not None and minimum<=measure<=120:
      g[label+"_weight"]+=w;g[label+("_hours_sum" if label=="actual" else "_sum")]+=w*measure
    if h is not None and 1<=h<=120 and inc is not None and inc>0:
     g["paid_weight"]+=w;g["paid_income_monthly"]+=w*inc;g["paid_hours"]+=w*h;g["paid_individual_hourly_sum"]+=w*inc/(h*52/12)
    if population["sample_n"]%100000==0:print(f"PNAD official parser: {population['sample_n']:,} records",flush=True)
 rows=list(grouped.values())
 save(PROV/"pnad_cells_rows.json",rows);save(PROV/"pnad_population_rows.json",[population])
 return build_pnad(rows,[population])

def official_rais():
 import openpyxl
 path=download(MTE_BASE+"4-tabelas_rais-2022.xlsx",PROV/"4-tabelas_rais-2022.xlsx")
 workbook=openpyxl.load_workbook(path,read_only=True,data_only=True)
 sheet=workbook["TABELA 2"]
 block=[]
 for idx in range(87,96):
  label=sheet.cell(idx,3).value;n=sheet.cell(idx,4).value
  if not isinstance(n,(int,float)) or "empregados" not in str(label):raise ValueError("Official RAIS size table layout changed")
  block.append({"code":str(idx-85),"label":str(label).replace("\u200b",""),"n_active_links":int(n),"source_cell":f"'TABELA 2'!C{idx}:D{idx}"})
 total=sum(r["n_active_links"] for r in block)
 if total!=52790864:raise ValueError("Official RAIS total disagrees with published 2022 stock")
 for r in block:r["share"]=r["n_active_links"]/total
 small=sum(r["n_active_links"] for r in block[:4])
 result={"status":"verified_official_fallback","source":"MTE RAIS 2022 official workbook, TABELA 2 C87:D95; corroborated by Executive Summary Table 6, page 9","source_url":MTE_BASE+"4-tabelas_rais-2022.xlsx","year":2022,"n_active_links":total,"small_le49_share_formal":small/total,"large_ge50_share_formal":1-small/total,"breakdown":block,"unit":"Employment links active 31 December 2022, not unique people","size_unit":"Establishment, not consolidated firm","universe":"All formal employment links (public/private and other legal types), all Brazil","informality_by_size":None,"informality_by_size_status":"Original 50%/20% are unverified hypotheses, not RAIS observations","contracted_hours_counts":None,"contracted_hours_status":"Microdata BigQuery blocked; no contracted-hours distribution inferred from PNAD","original_claims":{"small_share":.59,"subcategories_sum":.32,"source_verdict":"Neither matches the official 2022 active-links establishment-size table"},"pnad_comparability":"Do not merge these administrative establishment shares with all-worker PNAD informality without a universe bridge."}
 save(OUT/"rais_targets.json",result);save(FINAL/"rais_targets.json",result)
 return result

def main():
 parser=argparse.ArgumentParser(description=__doc__)
 parser.add_argument("--project",default="upa-research")
 parser.add_argument("--maximum-bytes-billed",type=int,default=20_000_000_000)
 parser.add_argument("--from-cache",action="store_true")
 parser.add_argument("--allow-official-fallback",action="store_true",help="Explicitly allow official IBGE/MTE sources if BigQuery is unavailable")
 parser.add_argument("--official-only",action="store_true",help="Use previously documented official fallback without repeating BigQuery authentication")
 args=parser.parse_args()
 for p in [OUT,FINAL,PROV]:p.mkdir(parents=True,exist_ok=True)
 jobs={**DICTS,"pnad_population":PNAD_POP_SQL,"pnad_cells":PNAD_SQL,"rais_cells":RAIS_SQL}
 for n,sql in jobs.items():(PROV/f"{n}.sql").write_text(sql,encoding="utf-8")
 route="bigquery"
 cached_official=args.from_cache and (OUT/"manifest.json").exists() and json.loads((OUT/"manifest.json").read_text(encoding="utf-8")).get("route")=="official_ibge_mte_fallback"
 try:
  if args.official_only:raise RuntimeError("Explicit official-only route selected; see bigquery_blocker.json")
  if args.from_cache:
   names=["pnad_cells","pnad_population"] if cached_official else jobs
   results={n:json.loads((PROV/f"{n}_rows.json").read_text(encoding="utf-8")) for n in names}
  else:
   from google.cloud import bigquery
   client=bigquery.Client(project=args.project)
   for table in ["basedosdados.br_ibge_pnadc.microdados","basedosdados.br_me_rais.microdados_vinculos"]:
    save(PROV/f"{table.split('.')[1]}_schema.json",client.get_table(table).to_api_repr())
   results={n:query(client,n,sql,args.maximum_bytes_billed) for n,sql in jobs.items()}
  pnad=build_pnad(results["pnad_cells"],results["pnad_population"])
  rais=official_rais() if cached_official else build_rais(results["rais_cells"],results["rais_dictionary"])
  if cached_official:
   route="official_ibge_mte_fallback"
   pnad["status"]="verified_official_fallback"
   pnad["metadata"].update({"source":"IBGE official public microdata (explicit fallback after BigQuery auth/access failure)","table":None,"file_vintage":IBGE_FILE,"dictionary_vintage":"Dicionario_e_input_20221031.zip","source_url":IBGE_BASE+"2024/"+IBGE_FILE,"recomputed_from":"Archived aggregate cells from this collector, not legacy model outputs"})
   save(OUT/"pnad_targets.json",pnad);save(FINAL/"pnad_targets.json",pnad)
 except Exception as exc:
  if not args.official_only:save(PROV/"bigquery_blocker.json",{"status":"blocked","project":args.project,"error":str(exc),"utc":dt.datetime.now(dt.timezone.utc).isoformat(),"action":"Reauthenticate the personal account and its application-default credentials for authorized project upa-research; corporate billing project is not used."})
  if not(args.allow_official_fallback or args.official_only):raise
  print(f"BigQuery unavailable: {type(exc).__name__}: {exc}\nUsing explicitly allowed official IBGE/MTE fallback, SAME 2024Q4.",flush=True)
  route="official_ibge_mte_fallback"
  pnad=official_pnad();rais=official_rais()
  pnad["status"]="verified_official_fallback"
  pnad["metadata"].update({"source":"IBGE official public microdata (explicit fallback after BigQuery auth/access failure)","table":None,"file_vintage":IBGE_FILE,"dictionary_vintage":"Dicionario_e_input_20221031.zip","source_url":IBGE_BASE+"2024/"+IBGE_FILE})
  save(OUT/"pnad_targets.json",pnad);save(FINAL/"pnad_targets.json",pnad)
 manifest={"status":"complete_with_explicit_bq_blocker" if route!="bigquery" else "complete","route":route,"pnad_period":"2024Q4","rais_year":2022,"utc":dt.datetime.now(dt.timezone.utc).isoformat(),"bigquery_billed_bytes":sum(json.loads(p.read_text())["total_bytes_billed"] or 0 for p in PROV.glob("*_job.json")),"digests":{str(p.relative_to(ROOT)):sha256_file(p) for p in [OUT/"pnad_targets.json",OUT/"rais_targets.json",PROV/"pnad_cells_rows.json",PROV/"pnad_population_rows.json"]}}
 save(OUT/"manifest.json",manifest);save(FINAL/"manifest.json",manifest)
 rows=json.loads((PROV/"pnad_cells_rows.json").read_text(encoding="utf-8"))
 old_inf=sum(r["occupied_weighted"] for r in rows if (r["position"] in (2,4,10) or (r["position"] in (8,9) and r["partner"]!=1)))
 audit={"old_V4017_classification_informality_rate_same_2024Q4":old_inf/pnad["national"]["occupied_weighted"],"correct_V4019_informality_rate":pnad["national"]["informality_rate"],"cnpj_misclassified_weighted":sum(r["occupied_weighted"] for r in rows if r["position"] in (8,9) and int(r["partner"]!=1)!=r["informal"]),"sample_notes":"Same raw2024Q4 sample: isolates variable coding error from data vintage changes","hours":pnad["national"]["formal_hours_proxy_cap44"],"private_employee_wage_bridge":pnad["national"]["private_employee_wage_bridge"]}
 save(OUT/"empirical_audit.json",audit)
 print(json.dumps({"pnad_informality":pnad["national"]["informality_rate"],"pnad_occupied":pnad["national"]["occupied_weighted"],"rais_small_share":rais["small_le49_share_formal"],"rais_active_links":rais["n_active_links"]},indent=2))

if __name__=="__main__":main()
